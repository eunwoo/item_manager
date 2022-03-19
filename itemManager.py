import sys

from PySide6.QtCore import QSize, Qt, QSortFilterProxyModel
from PySide6.QtGui import QAction, QIcon, QKeySequence, QColor, QBrush, QStandardItemModel, QStandardItem
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QLabel,
    QMainWindow,
    QStatusBar,
    QToolBar,
    QHBoxLayout,
    QVBoxLayout,
    QWidget,
    QTableWidget,
    QTableWidgetItem,
    QLineEdit,
    QDialog,
    QPushButton,
    QFileDialog,
    QHeaderView,
    QStyledItemDelegate,
    QStyle,
    QTabWidget,
    QMessageBox,
    QComboBox,
)
import xlsxwriter
from openpyxl import load_workbook
from data import Data
import json
import os.path

# Item Delegate
# https://pythonshowcase.com/question/background-color-of-the-particular-cell-is-not-changing-after-clicking-specific-cells-in-qtablewidget-pyqt5
# 
# 
# #
DATE_FILE = 'data.json'
APP_VERSION = 'Ver 8.0'

class NewItemDialog(QDialog):
    def __init__(self, title="새로운 아이템", name="", price="", stock="", parent=None):
        QDialog.__init__(self,parent)   # 또는 super().__init__(parent)

        mainLayout = QVBoxLayout()
        # 이름
        layout = QHBoxLayout()
        layout.addWidget(QLabel("이름"))
        self.name = QLineEdit(name)
        layout.addWidget(self.name)
        mainLayout.addLayout(layout)
        # 가격
        layout = QHBoxLayout()
        layout.addWidget(QLabel("가격"))
        self.price = QLineEdit(price)
        layout.addWidget(self.price)
        mainLayout.addLayout(layout)
        # 재고
        layout = QHBoxLayout()
        layout.addWidget(QLabel("재고"))
        self.stock = QLineEdit(stock)
        layout.addWidget(self.stock)
        mainLayout.addLayout(layout)

        # okButton, cancelButton 생성
        okButton = QPushButton("OK")
        cancelButton = QPushButton("Cancel")
        buttonLayout = QHBoxLayout()
        buttonLayout.addWidget(okButton)
        buttonLayout.addWidget(cancelButton)
        mainLayout.addLayout(buttonLayout)
        self.setLayout(mainLayout)
        self.setWindowTitle(title)

        okButton.clicked.connect(self.accept)     # accept() 슬롯에 연결
        cancelButton.clicked.connect(self.reject)  # reject() 슬롯에 연결

    def getResult(self):
        return (self.name, self.stock, self.price)
class DialogAsk(QDialog):
    def __init__(self,parent=None):
        QDialog.__init__(self,parent)   # 또는 super().__init__(parent)

        mainLayout = QVBoxLayout()
        mainLayout.addWidget(QLabel("삭제하시겠습니까?"))
        buttonLayout = QHBoxLayout()
        okButton = QPushButton("예")
        cancelButton = QPushButton("아니오")
        buttonLayout.addWidget(okButton)
        buttonLayout.addWidget(cancelButton)
        mainLayout.addLayout(buttonLayout)
        self.setLayout(mainLayout)
        self.setWindowTitle('확인해주세요')
        okButton.clicked.connect(self.accept)     # accept() 슬롯에 연결
        cancelButton.clicked.connect(self.reject)  # reject() 슬롯에 연결

class ItemTable(QTableWidget):
    def __init__(self, parent=None):
        QTableWidget.__init__(self, parent)
        
    def keyPressEvent(self, event):
        print('keypressed')
        if self.currentItem() == None:
            return
        print(self.currentItem().text())
        print(event.key())
        if event.key() == Qt.Key_Delete:
            dialog = DialogAsk()
            if dialog.exec():
                print('yes')
                indexes = []
                for selectionRange in self.selectedRanges():
                    indexes.extend(range(selectionRange.topRow(), selectionRange.bottomRow()+1))
                indexes.reverse()
                print(indexes)
                for i in indexes:
                    self.selectRow(i)
                    self.removeRow(self.currentRow())
            else:
                print('no')
        elif event.key() == Qt.Key_Right:
            row = self.currentRow()
            col = self.currentColumn()
            if col<2 and self.item(row, col+1):
                self.setCurrentCell(row, col+1)
        elif event.key() == Qt.Key_Left:
            row = self.currentRow()
            col = self.currentColumn()
            if col>0 and self.item(row, col-1):
                self.setCurrentCell(row, col-1)
        elif event.key() == Qt.Key_Up:
            row = self.currentRow()
            col = self.currentColumn()
            if row>0 and self.item(row-1, col):
                self.setCurrentCell(row-1, col)
                if self.item(row-1, col) == None:
                    self.setItem(row-1, col, QTableWidgetItem(""))
        elif event.key() == Qt.Key_Down:
            row = self.currentRow()
            col = self.currentColumn()
            if row<1000:
                self.setCurrentCell(row+1, col)
                if self.item(row+1, col) == None:
                    self.setItem(row+1, col, QTableWidgetItem(""))
        elif event.key() == Qt.Key_Return:
            print('enter')
            self.edit(self.currentIndex())
            


class MainWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()

        layout1 = QHBoxLayout()
        self.label = QLabel("검색")
        self.label.setAlignment(Qt.AlignCenter)
        layout1.addWidget(self.label)
        self.textbox = QLineEdit("")
        self.textbox.textChanged[str].connect(self.onChanged)
        layout1.addWidget(self.textbox)
        self.buttonFind = QPushButton("찾기")
        layout1.addWidget(self.buttonFind)
        self.buttonFind.clicked.connect(self.onClickFind)
        layout.addLayout(layout1)

        self.tableWidget = ItemTable(self)
        self.tableWidget.setRowCount(1000)
        self.tableWidget.setColumnCount(3)
        self.tableWidget.setColumnWidth(0, 160)
        self.tableWidget.setColumnWidth(1, 80)
        self.tableWidget.setColumnWidth(2, 80)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        itemA = QTableWidgetItem("이름")
        self.tableWidget.setHorizontalHeaderItem(0, itemA)
        itemA = QTableWidgetItem("가격")
        self.tableWidget.setHorizontalHeaderItem(1, itemA)
        itemA = QTableWidgetItem("재고")
        self.tableWidget.setHorizontalHeaderItem(2, itemA)

        # itemA = QTableWidgetItem("itemA")
        # self.tableWidget.setItem(0,1,itemA)

        # itemB = QTableWidgetItem()
        # itemB.setText("itemB")
        # itemB.setIcon(QIcon("dog.png"))
        # self.tableWidget.setItem(3,2,itemB)

        layout.addWidget(self.tableWidget)

        self.setLayout(layout)
    def Find(self, filter_text):
        for i in range(self.tableWidget.rowCount()):
            if self.tableWidget.item(i, 0) == None:
                break
            item = self.tableWidget.item( i, 0 )
            print(filter_text)
            print(item.text())
            if filter_text not in item.text():
                self.tableWidget.setRowHidden( i, True)
            else:
                self.tableWidget.setRowHidden( i, False)

    def onChanged(self, text):
        print('changed')
        print(text)
        self.Find(text)
    def onClickFind(self):
        print('find')
        # show all
        if self.textbox.text() == "":
            for i in range(self.tableWidget.rowCount()):
                if self.tableWidget.item(i, 0) == None:
                    break
                self.tableWidget.setRowHidden( i, False)
            return
        filter_text = self.textbox.text()
        self.Find(filter_text)
        return
        msgBox = QMessageBox()
        msgBox.setWindowTitle("알림") # 메세지창의 상단 제목
        msgBox.setWindowIcon(QIcon("exclamation-circle.png")) # 메세지창의 상단 icon 설정
        msgBox.setIcon(QMessageBox.Information) # 메세지창 내부에 표시될 아이콘
        msgBox.setText("검색 결과") # 메세지 제목
        msgBox.setInformativeText("없습니다") # 메세지 내용
        msgBox.setStandardButtons(QMessageBox.Yes) # 메세지창의 버튼
        msgBox.setDefaultButton(QMessageBox.Yes) # 포커스가 지정된 기본 버튼
        msgBox.exec() # 클릭한 버튼 결과 리턴

class PriceWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()

        self.tableWidget = ItemTable(self)
        self.tableWidget.setRowCount(1000)
        self.tableWidget.setColumnCount(3)
        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        itemA = QTableWidgetItem("가격")
        self.tableWidget.setHorizontalHeaderItem(0, itemA)
        itemA = QTableWidgetItem("아이템1")
        self.tableWidget.setHorizontalHeaderItem(1, itemA)
        itemA = QTableWidgetItem("아이템2")
        self.tableWidget.setHorizontalHeaderItem(2, itemA)

        # itemA = QTableWidgetItem("itemA")
        # self.tableWidget.setItem(0,1,itemA)

        # itemB = QTableWidgetItem()
        # itemB.setText("itemB")
        # itemB.setIcon(QIcon("dog.png"))
        # self.tableWidget.setItem(3,2,itemB)

        layout.addWidget(self.tableWidget)

        self.setLayout(layout)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setGeometry(100,100, 500, 600)
        self.setWindowTitle(''.join(["아이템 관리자 - ", APP_VERSION]))
        self.setWindowIcon(QIcon('money.ico'))
        self.mainWidget = MainWidget()
        self.priceWidget = PriceWidget()
        self.tab = QTabWidget()
        self.tab.addTab(self.mainWidget, "아이템 리스트")
        self.tab.addTab(self.priceWidget, "가격 매칭 테이블")

        self.setCentralWidget(self.tab)

        if os.path.isfile(DATE_FILE) == True:
            fp = open(DATE_FILE, 'r', encoding='euc-kr')
            try:
                data = json.load(fp)
                if 'items' in data: self.items = data['items']
                else: self.items = []
                if 'matching' in data: self.matching = data['matching']
                else: self.matching = []
                print(self.items)
                print(self.matching)
                self.LoadData()
                fp.close()
            except json.decoder.JSONDecodeError as e:
                self.LoadDefaultVariable()
        else:
            self.LoadDefaultVariable()


        toolbar = QToolBar("My main toolbar")
        toolbar.setIconSize(QSize(16, 16))
        self.addToolBar(toolbar)

        button_action = QAction(QIcon("new.png"), "새로만들기(&N)", self)
        button_action.setStatusTip("새로만들기")
        button_action.triggered.connect(self.onNew)
        # button_action.setCheckable(True)
        button_action.setShortcut(QKeySequence("Ctrl+N"))
        toolbar.addAction(button_action)

        toolbar.addSeparator()

        button_action_copy = QAction(QIcon("document-copy.png"), "복제하기(&U)", self)
        button_action_copy.setStatusTip("복제하기")
        button_action_copy.triggered.connect(self.onCopy)
        # button_action.setCheckable(True)
        button_action_copy.setShortcut(QKeySequence("Ctrl+U"))
        toolbar.addAction(button_action_copy)

        toolbar.addSeparator()

        button_action_toggleActivate = QAction(QIcon("cross-circle.png"), "활성화/비활성화(&E)", self)
        button_action_toggleActivate.setStatusTip("활성화/비활성화")
        button_action_toggleActivate.triggered.connect(self.onToggleActivate)
        # button_action2.setCheckable(True)
        button_action_toggleActivate.setShortcut(QKeySequence("Ctrl+E"))
        toolbar.addAction(button_action_toggleActivate)

        self.setStatusBar(QStatusBar(self))

        button_action_export1 = QAction(QIcon("document-export.png"), "내보내기1(&O)", self)
        button_action_export1.setStatusTip("활성아이템만 내보내기")
        button_action_export1.triggered.connect(self.onExport1)
        # button_action.setCheckable(True)
        button_action_export1.setShortcut(QKeySequence("Ctrl+O"))
        toolbar.addAction(button_action_export1)

        self.setStatusBar(QStatusBar(self))

        button_action_export2 = QAction(QIcon("document-export.png"), "내보내기2(&P)", self)
        button_action_export2.setStatusTip("전체 내보내기")
        button_action_export2.triggered.connect(self.onExport2)
        # button_action.setCheckable(True)
        button_action_export2.setShortcut(QKeySequence("Ctrl+P"))
        toolbar.addAction(button_action_export2)

        self.setStatusBar(QStatusBar(self))

        button_action_import = QAction(QIcon("document-import.png"), "불러오기(&I)", self)
        button_action_import.setStatusTip("불러오기")
        button_action_import.triggered.connect(self.onImport)
        # button_action.setCheckable(True)
        button_action_import.setShortcut(QKeySequence("Ctrl+I"))
        toolbar.addAction(button_action_import)

        self.setStatusBar(QStatusBar(self))

        button_action_up = QAction(QIcon("up.png"), "위로", self)
        button_action_up.setStatusTip("위로 이동")
        button_action_up.triggered.connect(self.onMoveUp)
        # button_action.setCheckable(True)
        button_action_up.setShortcut(QKeySequence("Ctrl+Q"))
        toolbar.addAction(button_action_up)

        self.setStatusBar(QStatusBar(self))

        button_action_down = QAction(QIcon("down.png"), "아래로", self)
        button_action_down.setStatusTip("아래로 이동")
        button_action_down.triggered.connect(self.onMoveDown)
        # button_action.setCheckable(True)
        button_action_down.setShortcut(QKeySequence("Ctrl+A"))
        toolbar.addAction(button_action_down)

        toolbar.addWidget(QLabel("기준값 x "))
        self.price_multiply = QLineEdit("1.0")
        self.price_multiply.setFixedWidth(50)
        toolbar.addWidget(self.price_multiply)
        self.combo_box = QComboBox()
        self.combo_box.addItem("그대로")
        self.combo_box.addItem("아이템1")
        self.combo_box.addItem("아이템2")
        toolbar.addWidget(self.combo_box)
        ############# 메뉴 ###############
        menu = self.menuBar()

        file_menu = menu.addMenu("파일(&F)")
        file_menu.addAction(button_action)
        file_menu.addAction(button_action_copy)
        file_menu.addAction(button_action_export1)
        file_menu.addAction(button_action_export2)
        file_menu.addAction(button_action_import)
        file_menu.addAction(button_action_toggleActivate)

    def LoadData(self):
        for i in range(len(self.items)):
            print(self.items[i])
            print(self.items[i][0])
            print(self.items[i][1])
            print(self.items[i][2])
            itema = QTableWidgetItem(self.items[i][0])
            self.mainWidget.tableWidget.setItem(i,0,itema)
            itema = QTableWidgetItem(str(self.items[i][1]))
            self.mainWidget.tableWidget.setItem(i,1,itema)
            itema = QTableWidgetItem(str(self.items[i][2]))
            self.mainWidget.tableWidget.setItem(i,2,itema)
        for i in range(len(self.matching)):
            print(self.matching[i])
            print(self.matching[i][0])
            print(self.matching[i][1])
            print(self.matching[i][2])
            itema = QTableWidgetItem(str(self.matching[i][0]))
            self.priceWidget.tableWidget.setItem(i,0,itema)
            itema = QTableWidgetItem(str(self.matching[i][1]))
            self.priceWidget.tableWidget.setItem(i,1,itema)
            itema = QTableWidgetItem(str(self.matching[i][2]))
            self.priceWidget.tableWidget.setItem(i,2,itema)
    def LoadDefaultVariable(self):
        print('LoadDefaultVariable')

    def closeEvent(self, e):
        print('close')
        fp = open(DATE_FILE, 'w', encoding='euc-kr')

        self.items = {}
        items = list()
        for i in range(self.mainWidget.tableWidget.rowCount()):
            if self.mainWidget.tableWidget.item(i, 0) == None:
                break
            if len(self.mainWidget.tableWidget.item(i, 0).text().strip()) == 0:   # 내용이 공백인 경우
                break
            one_item = [None]*3
            one_item[0] = self.mainWidget.tableWidget.item(i, 0).text()
            if self.mainWidget.tableWidget.item(i, 1) == None:
                one_item[1] = "0"
            else:
                one_item[1] = self.mainWidget.tableWidget.item(i, 1).text()
            if self.mainWidget.tableWidget.item(i, 2) == None:
                one_item[2] = "0"
            else:
                one_item[2] = self.mainWidget.tableWidget.item(i, 2).text()
            items.append(one_item)
        self.items['items'] = items

        items = list()
        for i in range(self.priceWidget.tableWidget.rowCount()):
            if self.priceWidget.tableWidget.item(i, 0) == None:
                break
            one_item = [None]*3
            one_item[0] = self.priceWidget.tableWidget.item(i, 0).text()
            one_item[1] = self.priceWidget.tableWidget.item(i, 1).text()
            one_item[2] = self.priceWidget.tableWidget.item(i, 2).text()
            items.append(one_item)
        self.items['matching'] = items

        json.dump(self.items, fp, indent = 2, ensure_ascii = False)
        fp.close()
        print('close')

    def DisableTableRow(self, tablewidget, row):
        self.TableToggleItemEditAttribute(tablewidget.item(row, 0))
        self.TableToggleItemEditAttribute(tablewidget.item(row, 1))
        self.TableToggleItemEditAttribute(tablewidget.item(row, 2))
        
    def TableToggleItemEditAttribute(self, item):
        if item.flags() & Qt.ItemIsEditable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setBackground(QColor(200,200,200))
            item.setForeground(QBrush(QColor(255,255,255)))
        else:
            item.setFlags(item.flags() | Qt.ItemIsEditable)
            item.setBackground(QColor(255,255,255))
            item.setForeground(QBrush(QColor(0,0,0)))

    def onToggleActivate(self, s):
        print("onToggleActivate", s)
        print(self.tab.currentIndex())
        if self.tab.currentIndex() == 0:
            indexes = []
            for selectionRange in self.mainWidget.tableWidget.selectedRanges():
                indexes.extend(range(selectionRange.topRow(), selectionRange.bottomRow()+1))
            for i in indexes:
                item = self.mainWidget.tableWidget.item(i, 0)
                if item == None:
                    return
                self.DisableTableRow(self.mainWidget.tableWidget, i)

        else:
            indexes = []
            for selectionRange in self.priceWidget.tableWidget.selectedRanges():
                indexes.extend(range(selectionRange.topRow(), selectionRange.bottomRow()+1))
            for i in indexes:
                item = self.priceWidget.tableWidget.item(i, 0)
                if item == None:
                    return
                self.DisableTableRow(self.priceWidget.tableWidget, i)

    def onNew(self, s):
        print("New", s)
        input_dialog = NewItemDialog("새로운 아이템")
        if input_dialog.exec():
            self.name, self.stock, self.price = input_dialog.getResult()
            print(''.join(["name = ", self.name.text()]))
            self.InsertItem(self.name.text(), self.price.text(), self.stock.text())
        else:
            print('b')

    def onCopy(self, s):
        print("Copy", s)
        row = self.mainWidget.tableWidget.currentRow()
        self.mainWidget.tableWidget.setCurrentCell(row, 1)
        price = self.mainWidget.tableWidget.currentItem().text()
        self.mainWidget.tableWidget.setCurrentCell(row, 2)
        stock = self.mainWidget.tableWidget.currentItem().text()
        input_dialog = NewItemDialog("아이템 복제","", price, stock)
        if input_dialog.exec():
            self.name, self.stock, self.price = input_dialog.getResult()
            print(''.join(["name = ", self.name.text()]))
            self.InsertItem(self.name.text(), self.price.text(), self.stock.text())
        else:
            print('b')

    def is_number(self, string):
        try:
            float(string)
            return True
        except ValueError:
            return False
    
    def ExportItemTable(self, workbook, export_option, is_only_editable = False):
        worksheet = workbook.add_worksheet('시트1')
        cell_format = workbook.add_format({'bold': False, 'font_color': 'black'})
        worksheet.write(0, 0, "품 명", cell_format)
        worksheet.write(0, 1, "가 격", cell_format)
        worksheet.write(0, 2, "재 고", cell_format)
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        worksheet.set_column(2, 2, 10)
        print_row = 0
        for i in range(self.mainWidget.tableWidget.rowCount()):
            if self.mainWidget.tableWidget.item(i, 0) == None:
                break
            if (not is_only_editable) or (self.mainWidget.tableWidget.item(i, 0).flags() & Qt.ItemIsEditable):
                worksheet.write_string(print_row+1, 0, self.mainWidget.tableWidget.item(i, 0).text(), cell_format)
                price_text = list()
                if self.is_number(self.mainWidget.tableWidget.item(i, 1).text()):
                    print('numeric test success')
                    # 기준값 적용
                    price_multiplied = float(self.mainWidget.tableWidget.item(i, 1).text())*float(self.price_multiply.text())
                    price_text.append(str(price_multiplied))
                    # 대체아이템
                    if export_option > 0:
                        price_text.append(self.GetEquivalentItem(float(price_text[0]), export_option))
                    worksheet.write(print_row+1, 1, " 또는 ".join(price_text), cell_format)
                else:
                    print('numeric test fail')
                if self.is_number(self.mainWidget.tableWidget.item(i, 2).text()):
                    worksheet.write_number(print_row+1, 2, float(self.mainWidget.tableWidget.item(i, 2).text()), cell_format)
                print_row = print_row + 1
    def ExportMatchingTable(self, workbook):
        worksheet = workbook.add_worksheet('시트2')
        cell_format = workbook.add_format({'bold': False, 'font_color': 'black'})
        worksheet.write(0, 0, "가격 범위", cell_format)
        worksheet.write(0, 1, "아이템 1", cell_format)
        worksheet.write(0, 2, "아이템 2", cell_format)
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(1, 1, 30)
        worksheet.set_column(2, 2, 30)
        print_row = 0
        for i in range(self.priceWidget.tableWidget.rowCount()):
            if self.priceWidget.tableWidget.item(i, 0) == None:
                break
            if self.priceWidget.tableWidget.item(i, 0).flags() & Qt.ItemIsEditable:
                worksheet.write_string(print_row+1, 0, self.priceWidget.tableWidget.item(i, 0).text(), cell_format)
                if self.priceWidget.tableWidget.item(i, 1):
                    worksheet.write_string(print_row+1, 1, self.priceWidget.tableWidget.item(i, 1).text(), cell_format)
                if self.priceWidget.tableWidget.item(i, 2):
                    worksheet.write_string(print_row+1, 2, self.priceWidget.tableWidget.item(i, 2).text(), cell_format)
                print_row = print_row + 1

    def onExport1(self, s):#활성화된 내용만 내보내기
        print("Export", s)
        filename, selectgedFilter = QFileDialog.getSaveFileName(self, "Save Excel File", ".",
                                                        "Excel File (*.xls *.xlsx)")
        if filename != "":
            export_option = self.combo_box.currentIndex()
            workbook = xlsxwriter.Workbook(filename)
            self.ExportItemTable(workbook, export_option, True)
            self.ExportMatchingTable(workbook)
            workbook.close()

    def onExport2(self, s):#전체 내용 내보내기
        print("Export", s)
        filename, selectgedFilter = QFileDialog.getSaveFileName(self, "Save Excel File", ".",
                                                        "Excel File (*.xls *.xlsx)")
        if filename != "":
            export_option = self.combo_box.currentIndex()
            workbook = xlsxwriter.Workbook(filename)
            self.ExportItemTable(workbook, export_option, False)
            self.ExportMatchingTable(workbook)
            workbook.close()

    def GetEquivalentItem(self, price, option):
        item_text = ""
        for i in range(self.priceWidget.tableWidget.rowCount()):
            if self.priceWidget.tableWidget.item(i, option) == None:
                break
            price_range_str = self.priceWidget.tableWidget.item(i, 0).text()
            price_range = price_range_str.split('-')
            print(price_range)
            if price >= float(price_range[0].strip()) and price <= float(price_range[1].strip()):
                item_text = self.priceWidget.tableWidget.item(i, option).text()
                break
        return item_text

    def onImport(self, s):
        print("Import", s)
        filename, selectgedFilter = QFileDialog.getOpenFileName(self, "Open Excel File", ".",
                                                        "Excel File (*.xls *.xlsx)")
        if filename != "":
            print('a')
            print(filename)
            wb = load_workbook(filename)
            if '시트1' in wb:
                sheet_ranges = wb['시트1']
            elif 'Sheet1' in wb:
                sheet_ranges = wb['Sheet1']
            else:
                return
            for i in range(1000):
                cellName = ''.join(['A', str(i+2)])
                name = sheet_ranges[cellName].value
                if name == None:
                    break
                print(name)
                print(type(name))
                cellPrice = ''.join(['B', str(i+2)])
                price = sheet_ranges[cellPrice].value
                if price == None:
                    price = ""
                print(price)
                print(type(price))
                cellStock = ''.join(['C', str(i+2)])
                stock = sheet_ranges[cellStock].value
                if stock == None:
                    stock = ""
                print(stock)
                print(type(stock))
                # data = Data(name, price, stock)
                self.AddItem(name, str(price), str(stock))
            # self.mainWidget.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            # self.mainWidget.tableWidget.resizeColumnsToContents()
            if '시트2' in wb:
                sheet_ranges = wb['시트2']
            elif 'Sheet2' in wb:
                sheet_ranges = wb['Sheet2']
            else:
                return
            print("load price table")
            for i in range(1000):
                cell = ''.join(['A', str(i+2)])
                price = sheet_ranges[cell].value
                print(price)
                if price == None:
                    break
                cell = ''.join(['B', str(i+2)])
                item1 = sheet_ranges[cell].value
                cell = ''.join(['C', str(i+2)])
                item2 = sheet_ranges[cell].value
                self.AddTableItem(str(price), str(item1), str(item2))

        else:
            print('b')

    def AddItem(self, name, price, stock):
        row = self.GetEmptyRow(self.mainWidget.tableWidget)
        itema = QTableWidgetItem(name)
        self.mainWidget.tableWidget.setItem(row,0,itema)
        itema = QTableWidgetItem(price)
        self.mainWidget.tableWidget.setItem(row,1,itema)
        itema = QTableWidgetItem(stock)
        self.mainWidget.tableWidget.setItem(row,2,itema)

    def InsertItem(self, name, price, stock):
        row = self.mainWidget.tableWidget.currentRow() + 1
        self.mainWidget.tableWidget.insertRow(row)
        itema = QTableWidgetItem(name)
        self.mainWidget.tableWidget.setItem(row,0,itema)
        itema = QTableWidgetItem(price)
        self.mainWidget.tableWidget.setItem(row,1,itema)
        itema = QTableWidgetItem(stock)
        self.mainWidget.tableWidget.setItem(row,2,itema)
        self.mainWidget.tableWidget.selectRow(row)

    def AddTableItem(self, price, item1, item2):
        print('add table item')
        row = self.GetEmptyRow(self.priceWidget.tableWidget)
        itema = QTableWidgetItem(price)
        self.priceWidget.tableWidget.setItem(row,0,itema)
        itema = QTableWidgetItem(item1)
        self.priceWidget.tableWidget.setItem(row,1,itema)
        itema = QTableWidgetItem(item2)
        self.priceWidget.tableWidget.setItem(row,2,itema)

    def GetEmptyRow(self, tablewidget):
        for i in range(tablewidget.rowCount()):
            if tablewidget.item(i, 0) == None:
                return i
        if i == tablewidget.rowCount() - 1:
            tablewidget.insertRow(i+1)
            return i+1

    def SwapTableItem(self, tablewidget, rowSrc, rowDest):
        itemSrc = tablewidget.takeItem(rowSrc, 0)
        itemDst = tablewidget.takeItem(rowDest, 0)
        tablewidget.setItem(rowSrc, 0, itemDst)
        tablewidget.setItem(rowDest, 0, itemSrc)
        itemSrc = tablewidget.takeItem(rowSrc, 1)
        itemDst = tablewidget.takeItem(rowDest, 1)
        tablewidget.setItem(rowSrc, 1, itemDst)
        tablewidget.setItem(rowDest, 1, itemSrc)
        itemSrc = tablewidget.takeItem(rowSrc, 2)
        itemDst = tablewidget.takeItem(rowDest, 2)
        tablewidget.setItem(rowSrc, 2, itemDst)
        tablewidget.setItem(rowDest, 2, itemSrc)

    def onMoveUp(self, s):
        print("up", s)
        row = self.mainWidget.tableWidget.currentRow()
        if row == 0:
            return
        self.SwapTableItem(self.mainWidget.tableWidget, row, row-1)
        self.mainWidget.tableWidget.selectRow(row - 1)

    def onMoveDown(self, s):
        print("down", s)
        row = self.mainWidget.tableWidget.currentRow()
        if row == self.mainWidget.tableWidget.rowCount() - 1:
            return
        self.SwapTableItem(self.mainWidget.tableWidget, row, row + 1)
        self.mainWidget.tableWidget.selectRow(row + 1)

    # def MakeCSV(self):
    #     first = 0
    #     index1 = 0
    #     for dl in self.logs:
    #         print(dl.commit_hash)
    #         if first == 0:
    #             first = 1
    #             df = pd.DataFrame({'commit hash': dl.commit_hash}, index = [index1])
    #             print(index1)
    #         else:
    #             df = pd.concat([df, pd.DataFrame({'commit hash': dl.commit_hash}, index = [index1])])
    #             print(index1)
    #         index1 = index1 + 1
    #     print(df)
    #     filename_csv = target_folder+'/log.csv'
    #     df.to_csv(filename_csv)


app = QApplication(sys.argv)

window = MainWindow()
window.show()

app.exec()